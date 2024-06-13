'''
读取表格信息
'''
from openpyxl import load_workbook
from settings import BASE_DIR

class ReadData:
    # 定义类方法，方便调用
    @classmethod
    def readExcel(cls, filename, sheetName):
        '''

        :param filename: Excel文件名字
        :param sheetName: sheet的名字
        :return: 返回的是读取出来列表中的所有数据
        '''
        #这里调式的时候路径要写两个点，后面项目执行的时候读取数据要写一个点
        # 读取已经存在的表格(把在外面创建的表格，通过复制粘贴的方法放到这个目录下，如mms_eg.xlsx)
        wb = load_workbook(f'{BASE_DIR}./data/{filename}.xlsx')

        # 指定读取表格的哪一个sheet
        ws = wb[sheetName]
        # 获取最大行
        rows = ws.max_row
        # 获取最大列
        cols = ws.max_column
        # 定义大列表，保存值[[],[],[],[]]
        rows_list = []

        # 利用嵌套for循环完成读取
        # 从第二行开始读,取到最后一行。rang是左开右闭
        for r in range(2, rows + 1):
            # 定义一个列表，用来保存取出的每一行信息
            rows_value = []
            for c in range(1, cols + 1):
                # 根据行和列取值，读取信息
                i = ws.cell(r, c).value
                s = str(i).strip()
                # 给列表里面添加值
                rows_value.append(s)
            rows_list.append(rows_value)

        return rows_list


#只在当前页面执行脚本，进行检查代码准确性
if __name__=='__main__':
    r = ReadData.readExcel('mms_eg','查询顾客信息')
    print(r)